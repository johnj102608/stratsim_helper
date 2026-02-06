import sys
from pathlib import Path
import json
import re
import pandas as pd
from openpyxl import load_workbook

# ======================
# BASE
# ======================
#BASE_DIR = Path(__file__).parent
#BASE_DIR = Path(sys.executable).parent
def get_app_dir():
    if getattr(sys, 'frozen', False):
        # Running as a PyInstaller exe
        return Path(sys.executable).parent
    else:
        # Running as a normal Python script
        return Path(__file__).parent

BASE_DIR = get_app_dir()
# ======================
# JSON FILES (next to your script)
# ======================
CONFIG_PATH = BASE_DIR / "config.json"
ALIASES_PATH = BASE_DIR / "metric_aliases.json"

# ======================
# DEFAULT CONFIG (used if config.json missing)
# ======================
DEFAULT_CONFIG = {
    "dashboard_template_name": "StratSim_Dashboard_2025-FL_Section01.xlsx",
    "output_dashboard_name": "StratSim_Dashboard_UPDATED.xlsx",
    "input_prefix": "Competition - Financial Summary - Year ",
    "financial_sheet_prefix": "Financial Details for ",
    "firms": list("ABCDEFG"),
    "year_sheet_prefix": "Year ",
    "firm_prefix": "FIRM ",
    "scan_max_rows": 250,
    "scan_max_cols": 30,
    "look_right_max": 4
}


# ======================
# LOAD JSONS
# ======================
def load_config(path: Path) -> dict:
    """
    Load config.json with defaults + light validation.
    """
    cfg = DEFAULT_CONFIG.copy()

    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            user_cfg = json.load(f)
        if not isinstance(user_cfg, dict):
            raise ValueError("config.json must be a JSON object (dictionary).")
        cfg.update(user_cfg)
    else:
        print(f"Warning: config.json not found at {path}. Using defaults.")

    # normalize + validate
    if not isinstance(cfg.get("firms"), list) or not cfg["firms"]:
        raise ValueError("config.json: 'firms' must be a non-empty list like ['A','B',...].")
    cfg["firms"] = [str(x).strip().upper() for x in cfg["firms"]]

    for k in ("scan_max_rows", "scan_max_cols", "look_right_max"):
        try:
            cfg[k] = int(cfg[k])
        except Exception as e:
            raise ValueError(f"config.json: '{k}' must be an integer.") from e

    for k in (
        "dashboard_template_name",
        "output_dashboard_name",
        "input_prefix",
        "financial_sheet_prefix",
        "year_sheet_prefix",
    ):
        if not isinstance(cfg.get(k), str) or not cfg[k].strip():
            raise ValueError(f"config.json: '{k}' must be a non-empty string.")

    return cfg


def load_metric_aliases(path: Path) -> dict[str, str]:
    """
    Load metric_aliases.json mapping:
      { "starting inventory": "beg. inventory", ... }

    Keys/values normalized to lowercase+strip.
    Missing file -> empty mapping.
    """
    if not path.exists():
        print(f"Note: metric_aliases.json not found at {path}. No aliasing will be applied.")
        return {}

    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    if not isinstance(raw, dict):
        raise ValueError("metric_aliases.json must be a JSON object (dictionary).")

    out: dict[str, str] = {}
    for k, v in raw.items():
        kk = str(k).strip().lower()
        vv = str(v).strip().lower()
        if kk and vv:
            out[kk] = vv
    return out


def canonical_metric(metric: str, aliases: dict[str, str]) -> str:
    m = str(metric).strip().lower()
    return aliases.get(m, m)


# ======================
# READ CONFIGS
# ======================
cfg = load_config(CONFIG_PATH)
metric_aliases = load_metric_aliases(ALIASES_PATH)

DASHBOARD_TEMPLATE_PATH = BASE_DIR / cfg["dashboard_template_name"]
OUTPUT_DASHBOARD_PATH = BASE_DIR / cfg["output_dashboard_name"]

INPUT_PREFIX = cfg["input_prefix"]
FINANCIAL_SHEET_PREFIX = cfg["financial_sheet_prefix"]
FIRMS = cfg["firms"]
YEAR_SHEET_PREFIX = cfg["year_sheet_prefix"]
FIRM_PREFIX = cfg["firm_prefix"]

SCAN_MAX_ROWS = cfg["scan_max_rows"]
SCAN_MAX_COLS = cfg["scan_max_cols"]
LOOK_RIGHT_MAX = cfg["look_right_max"]




# ----------------------
# File helpers
# ----------------------
def parse_round_num(filename: str) -> int:
    """
    Extracts year number from:
      "Competition - Financial Summary - Year 2.xlsx" -> 2
    """
    name = filename.lower()
    name = name.replace(INPUT_PREFIX.lower(), "")
    name = name.replace(".xlsx", "")
    return int(name.strip())


def list_round_files() -> list[Path]:
    """Only .xlsx files in the folder that start with INPUT_PREFIX."""
    files: list[Path] = []
    for p in BASE_DIR.iterdir():
        if not p.is_file():
            continue
        if p.suffix.lower() != ".xlsx":
            continue
        if not p.name.lower().startswith(INPUT_PREFIX.lower()):
            continue
        files.append(p)
    return sorted(files, key=lambda p: parse_round_num(p.name))


def firm_sheet_name(letter: str) -> str:
    return f"{FINANCIAL_SHEET_PREFIX}{letter}"


def firm_label(letter: str) -> str:
    return f"{FIRM_PREFIX}{letter}"


# ----------------------
# Input parsing (scan-all firm sheets)
# ----------------------
_num_re = re.compile(r"^\(?-?\$?\d[\d,]*\.?\d*\)?%?$")


def is_numberish(x: str) -> bool:
    x = str(x).strip()
    if x == "":
        return False
    return bool(_num_re.match(x))


def to_number(x: str) -> float | None:
    s = str(x).strip()
    if s == "":
        return None

    s = s.replace(",", "").replace("$", "")

    # parentheses negatives: (123) -> -123
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]

    # percent: keep as numeric (e.g., "12.3%")->12.3
    s = s.replace("%", "")

    try:
        return float(s)
    except ValueError:
        return None


def is_label_cell(x: str) -> bool:
    """
    Candidate metric label if:
      - non-empty
      - does NOT look numeric
    """
    s = str(x).strip()
    if s == "":
        return False
    if is_numberish(s):
        return False
    return True


def read_sheet_raw(input_path: Path, sheet_name: str) -> pd.DataFrame:
    """
    Read the sheet as raw values and normalize whitespace.
    (pandas 2.x: DataFrame.map is element-wise)
    """
    df = pd.read_excel(input_path, sheet_name=sheet_name, header=None, dtype=object).fillna("")
    df = df.map(lambda v: str(v).strip())
    return df


def scan_pairs_in_df(df: pd.DataFrame, look_right_max: int = LOOK_RIGHT_MAX) -> pd.DataFrame:
    """
    Scan a DataFrame top->bottom, left->right, extracting:
      label -> nearest numeric value to its right (within look_right_max columns)

    Supports rows like:
      name value
      name1 value1 name2 value2

    Ignores rows with no numeric values at all (section headers).
    """
    records: list[dict[str, object]] = []

    for r in range(df.shape[0]):
        row = [str(v).strip() for v in df.iloc[r].tolist()]

        if all(v == "" for v in row):
            continue

        if not any(to_number(v) is not None for v in row):
            continue

        c = 0
        while c < len(row):
            cell = row[c]
            if is_label_cell(cell):
                found_val = None
                found_at = None

                for k in range(1, look_right_max + 1):
                    if c + k >= len(row):
                        break
                    num = to_number(row[c + k])
                    if num is not None:
                        found_val = num
                        found_at = c + k
                        break

                if found_val is not None:
                    records.append({"metric": cell.strip(), "value": float(found_val)})
                    c = found_at + 1
                    continue

            c += 1

    out = pd.DataFrame(records)
    if out.empty:
        return pd.DataFrame(columns=["metric", "value"])

    out["metric"] = out["metric"].astype(str).str.strip()
    out = out[out["metric"] != ""]
    return out


def read_firm_details_long(input_path: Path) -> pd.DataFrame:
    """
    Read all firm sheets and return long df:
      metric | firm | value
    using the scan-all algorithm (no reliance on block titles).
    Applies metric alias mapping (input -> dashboard name) if provided.
    """
    xls = pd.ExcelFile(input_path)
    rows: list[pd.DataFrame] = []

    for f in FIRMS:
        sheet = firm_sheet_name(f)
        if sheet not in xls.sheet_names:
            continue

        raw = read_sheet_raw(input_path, sheet)
        slab = raw.iloc[:SCAN_MAX_ROWS, :SCAN_MAX_COLS].copy()

        pairs = scan_pairs_in_df(slab, look_right_max=LOOK_RIGHT_MAX)
        if pairs.empty:
            print(f"  Warning: extracted 0 metric/value pairs from '{sheet}'")
            continue

        pairs["firm"] = firm_label(f)
        rows.append(pairs)

    if not rows:
        return pd.DataFrame(columns=["metric", "firm", "value"])

    out = pd.concat(rows, ignore_index=True)

    out["metric"] = out["metric"].astype(str).str.strip().str.lower()
    if metric_aliases:
        out["metric"] = out["metric"].map(lambda m: canonical_metric(m, metric_aliases))

    out["firm"] = out["firm"].astype(str).str.strip().str.upper()
    out["value"] = pd.to_numeric(out["value"], errors="coerce")
    out = out.dropna(subset=["metric", "firm", "value"])
    return out


# ----------------------
# Output writing (whole-sheet writer, no block titles, auto metric column)
# ----------------------
def find_firm_header_row(ws) -> int:
    """
    Find the row that looks most like the firm header row by counting
    how many cells start with 'FIRM'. Pick the max.
    """
    best_row = None
    best_count = 0

    for r in range(1, ws.max_row + 1):
        count = 0
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().upper().startswith(FIRM_PREFIX):
                count += 1
        if count > best_count:
            best_count = count
            best_row = r

    if best_row is None or best_count == 0:
        raise ValueError("Could not find a firm header row in this Year sheet.")
    return best_row


def build_firm_to_col(ws, header_row: int) -> dict[str, int]:
    """
    Map 'FIRM A'.. to columns in the header row.
    Uses the FIRST CONTIGUOUS BLOCK of firm headers to avoid hidden/duplicate sections.
    """
    firm_to_col: dict[str, int] = {}
    started = False

    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        name = v.strip().upper() if isinstance(v, str) else ""
        is_firm = name.startswith(FIRM_PREFIX)

        if is_firm and not started:
            started = True

        if started and not is_firm:
            break

        if is_firm and name not in firm_to_col:
            firm_to_col[name] = c

    if not firm_to_col:
        raise ValueError("No firm columns found on detected firm header row.")
    return firm_to_col


def _looks_numeric_string(s: str) -> bool:
    """
    Very forgiving numeric string check used for output heuristics.
    """
    s = s.strip()
    if not s:
        return False
    s2 = s.replace(",", "").replace("$", "").replace("%", "")
    if s2.startswith("(") and s2.endswith(")"):
        s2 = "-" + s2[1:-1]
    try:
        float(s2)
        return True
    except Exception:
        return False


def find_metric_column(ws, firm_header_row: int) -> int:
    """
    Auto-detect the metric-name column in the Year sheet.

    Heuristic:
      - choose the column with the most non-numeric text cells
      - exclude the firm header row
      - exclude columns that are firm columns (cells like "FIRM A" on the firm header row)
    """
    best_col = None
    best_score = 0

    for c in range(1, ws.max_column + 1):
        # Skip firm columns
        hv = ws.cell(firm_header_row, c).value
        if isinstance(hv, str) and hv.strip().upper().startswith(FIRM_PREFIX):
            continue

        score = 0
        for r in range(1, ws.max_row + 1):
            if r == firm_header_row:
                continue

            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            s = v.strip()
            if not s:
                continue

            # ignore firm-like strings
            if s.upper().startswith(FIRM_PREFIX):
                continue

            # ignore numeric-looking strings
            if _looks_numeric_string(s):
                continue

            score += 1

        if score > best_score:
            best_score = score
            best_col = c

    if best_col is None or best_score == 0:
        raise ValueError("Could not automatically determine the metric column in the Year sheet.")
    return best_col


def build_metric_to_row_whole_sheet(ws, metric_col: int) -> dict[str, int]:
    """
    Build mapping: metric label -> row index by scanning metric_col across the whole Year sheet.
    First occurrence wins (good for handoff; avoids unexpected overwrites).
    """
    metric_to_row: dict[str, int] = {}

    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, metric_col).value
        if isinstance(v, str) and v.strip():
            key = v.strip().lower()
            if key not in metric_to_row:
                metric_to_row[key] = r

    if not metric_to_row:
        raise ValueError("Could not find any metric labels in the detected metric column.")
    return metric_to_row


def fill_year_sheet(ws, details_long: pd.DataFrame) -> None:
    """
    Write all matching metric/firm values into the Year sheet.

    It auto-detects:
      - the firm header row
      - the metric column

    Requirements:
      - details_long has columns: metric (lower), firm (upper), value (float)
      - Year sheet has a firm header row with cells like 'FIRM A'...
      - Year sheet has a metric-name column with many text labels
    """
    header_row = find_firm_header_row(ws)
    firm_to_col = build_firm_to_col(ws, header_row)

    metric_col = find_metric_column(ws, header_row)
    metric_to_row = build_metric_to_row_whole_sheet(ws, metric_col=metric_col)

    written = 0
    for _, rec in details_long.iterrows():
        metric = str(rec["metric"]).strip().lower()
        firm = str(rec["firm"]).strip().upper()
        value = float(rec["value"])

        if metric in metric_to_row and firm in firm_to_col:
            ws.cell(metric_to_row[metric], firm_to_col[firm]).value = value
            written += 1

    #print(f"  Detected metric column: {metric_col}")
    print(f"  Wrote {written} cells")


def pause_if_exe():
    if getattr(sys, 'frozen', False):
        input("\nPress Enter to exit...")


# ----------------------
# Main
# ----------------------
def main():
    round_files = list_round_files()
    if not round_files:
        raise SystemExit("No round Excel files found.")

    print("Found round files:")
    for f in round_files:
        print(f" - {f.name} (year {parse_round_num(f.name)})")

    if not DASHBOARD_TEMPLATE_PATH.exists():
        raise SystemExit(f"Dashboard template not found: {DASHBOARD_TEMPLATE_PATH}")

    wb = load_workbook(DASHBOARD_TEMPLATE_PATH)

    for f in round_files:
        yr = parse_round_num(f.name)
        year_sheet = f"{YEAR_SHEET_PREFIX}{yr}"
        if year_sheet not in wb.sheetnames:
            print(f"Skipping {f.name}: no sheet named '{year_sheet}' in dashboard template.")
            continue

        print(f"\nProcessing {f.name} -> {year_sheet}")
        ws = wb[year_sheet]

        details_long = read_firm_details_long(f)
        fill_year_sheet(ws, details_long)

    wb.save(OUTPUT_DASHBOARD_PATH)
    print(f"\nSaved updated dashboard to: {OUTPUT_DASHBOARD_PATH}")
    pause_if_exe()


if __name__ == "__main__":
    main()
